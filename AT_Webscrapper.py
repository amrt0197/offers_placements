import boto3
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import range_boundaries
from io import BytesIO
from urllib.parse import urlparse
import json
import logging
import re
import time
from datetime import datetime, timezone
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
 
#from requests_html import HTMLSession
#working on webscrapper
 
# Configure logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)
 
DEFAULT_TIMEOUT = 20
MAX_WORKERS = 10
VALID_REGIONS = {"US", "UK"}
 
session = boto3.session.Session()
s3_client = session.client('s3')
sqs = boto3.client('sqs')
secrets_client = session.client('secretsmanager', region_name='us-east-2')
 
@lru_cache(maxsize=1)
def get_secret():
    secret_name = "cj-bpa-webscrapper"
    response = secrets_client.get_secret_value(SecretId=secret_name)
    return json.loads(response['SecretString'])
 
def extract_domain_name(url):
    try:
        hostname = urlparse(url if url.startswith('http') else 'http://' + url).hostname or ''
        parts = hostname.split('.')
        if len(parts) > 2 and parts[-2] in ('co', 'com', 'org', 'net'):
            return parts[-3]
        return parts[-2] if len(parts) >= 2 else hostname
    except:
        return 'unknown_domain'
 
#def sanitize_publisher_name(name):
    #sanitized = re.sub(r'[^\w\s-]', '', name).strip()
    #return (sanitized[0].upper() + sanitized[1:] if sanitized else 'UnknownPublisher').replace(' ', '_') + ".xlsx"
 
def write_excel_to_s3(data, bucket, key, headers):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in data:
        ws.append(row)
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    s3_client.put_object(Bucket=bucket, Key=key, Body=output_stream.getvalue())
    logger.info(f"Uploaded to s3://{bucket}/{key}")
 
def scrape_url(url, selector):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9',
        'Accept-Language': 'en-US,en;q=0.5'
    }
    try:
        session = requests.Session()
        #session = HTMLSession()
        response = session.get(url, headers=headers, timeout=DEFAULT_TIMEOUT)
        #response.html.render(timeout=DEFAULT_TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        #soup = BeautifulSoup(response.html.html, 'html.parser')
        elements = soup.select(selector)
        texts = [el.get_text(strip=True) for el in elements if el.get_text(strip=True)]
 
        if texts:
            return url, texts, None
        else:
            element_html = [str(el) for el in elements]
            logger.warning(f"[DEBUG] Found element at {url}, but it's empty. HTML snippet: {element_html[:1]}")
            return url, None, "Element found but empty"
 
    except requests.HTTPError as http_err:
        return url, None, f"HTTP error: {http_err.response.status_code} {http_err.response.reason}"
    except Exception as e:
        return url, None, f"Error: {str(e)}"
 
 
def process_excel_file(bucket, key, message_id, publisher_name, region, offer_type, unique_id_base):
    response = s3_client.get_object(Bucket=bucket, Key=key)
    workbook = load_workbook(filename=BytesIO(response['Body'].read()), data_only=False)
    sheet = workbook.active
    table = sheet.tables.get('Table1')
    if not table:
        logger.error(f"Table1 not found in {key}")
        return 0, 0
 
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    tasks = []
    for row in sheet.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        tasks.append((str(row[0]).strip(), str(row[1]).strip()))
 
    domain_data = defaultdict(lambda: {'success': [], 'error': []})
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
 
    # Track unique URLs and assign incremental base IDs
    unique_url_map = {}
    url_id_counter = int(unique_id_base)
    global_row_counter = 1
 
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_url = {executor.submit(scrape_url, url, sel): (url, sel) for url, sel in tasks}
        for future in as_completed(future_to_url):
            url, sel = future_to_url[future]
            try:
                result_url, results, error = future.result()
                domain = extract_domain_name(result_url)
                clean_url = result_url.strip().lower()
 
                if clean_url not in unique_url_map:
                    unique_url_map[clean_url] = str(url_id_counter)
                    url_id_counter += 1
 
                base_id = unique_url_map[clean_url]
 
                if error:
                    order_id = f"{base_id}-{global_row_counter}"
                    domain_data[domain]['error'].append([order_id, result_url, error])
                    global_row_counter += 1
                else:
                    for r in results:
                        order_id = f"{base_id}-{global_row_counter}"
                        domain_data[domain]['success'].append([order_id, result_url, r])
                        global_row_counter += 1
 
            except Exception as e:
                domain = extract_domain_name(url)
                clean_url = url.strip().lower()
 
                if clean_url not in unique_url_map:
                    unique_url_map[clean_url] = str(url_id_counter)
                    url_id_counter += 1
 
                base_id = unique_url_map[clean_url]
                order_id = f"{base_id}-{global_row_counter}"
                domain_data[domain]['error'].append([order_id, url, f"Unhandled Exception: {str(e)}"])
                global_row_counter += 1
 
    success_count = error_count = 0
    for domain, data in domain_data.items():
        if data['success']:
            key_success = f"Webscrapper Output/{region}/{offer_type}/{publisher_name}/{publisher_name}_{date_str}.xlsx"
            write_excel_to_s3(data['success'], bucket, key_success, ['Web Scraper Order ID', 'URL', 'Scraped Text'])
            success_count += 1
        if data['error']:
            key_error = f"Webscrapper Output/error_files/{region}/{offer_type}/{publisher_name}/{publisher_name}_{date_str}_error_file.xlsx"
            write_excel_to_s3(data['error'], bucket, key_error, ['Web Scraper Order ID', 'URL', 'Status/Error'])
            error_count += 1
 
    return success_count, error_count
 
 
def lambda_handler(event, context):
 
    processed_count = failed_count = 0
    try:
        secrets = get_secret()
        bucket_name = secrets.get('webscrapper_s3_bucket_name')
        queue_url_sqs = secrets.get('queue_url_feed_webscrapper')
 
        for record in event.get('Records', []):
            message_id = record.get('messageId', 'UnknownMessageID')
            receipt_handle = record.get('receiptHandle')
            try:
                message_data = json.loads(record['body'])
                publisher_name = message_data.get('Publisher Name')
                region = message_data.get('region_of_origin', '').upper()
                offer_type = message_data.get('Offers/Cashback', '')
                unique_id_base = message_data.get('unique_id')
 
                if not all([publisher_name, region, offer_type]):
                    failed_count += 1
                    continue
                if region not in VALID_REGIONS:
                    failed_count += 1
                    continue
 
                #formatted_name = sanitize_publisher_name(publisher_name)
                input_key = f"Input Files/{region}/{offer_type}/{publisher_name}.xlsx"
                success_count, error_count = process_excel_file(bucket_name, input_key, message_id,publisher_name,region,offer_type,unique_id_base)
                if success_count > 0 or error_count > 0:
                    processed_count += 1
                    # Delete message only if successfully processed
                    sqs.delete_message(QueueUrl=queue_url_sqs, ReceiptHandle=receipt_handle)
                else:
                    failed_count += 1
            except Exception as e:
                logger.error(f"Error processing message {message_id}: {str(e)}")
                failed_count += 1
 
        return {
            'statusCode': 200 if processed_count > 0 else 500,
            'body': json.dumps(f"Processed: {processed_count}, Failed: {failed_count}")
        }
 
    except Exception as e:
        return {
            'statusCode': 500,
            'body': json.dumps(f"Critical error: {str(e)}")
        }